"""Helpers for creating and updating the Excel workbook."""

from __future__ import annotations

import logging
from pathlib import Path
from threading import Lock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import COLUNAS


log = logging.getLogger(__name__)


SHEET_DADOS = "Chamados"
SHEET_RESUMO = "Resumo"

COR_CABECALHO_BG = "1F3864"
COR_CABECALHO_FG = "FFFFFF"
COR_LINHA_PAR = "DCE6F1"
COR_ALTA = "FF4444"
COR_MEDIA = "FFA500"
COR_BAIXA = "70AD47"
COR_ABERTO = "FFC000"
COR_RESOLVIDO = "70AD47"

PRIORIDADE_CORES = {
    "Alta": COR_ALTA,
    "Média": COR_MEDIA,
    "Baixa": COR_BAIXA,
}

STATUS_CORES = {
    "Aberto": COR_ABERTO,
    "Em andamento": "4472C4",
    "Resolvido": COR_RESOLVIDO,
    "Cancelado": "A5A5A5",
}


def borda_fina() -> Border:
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


class ExcelHandler:
    def __init__(self, caminho: str):
        self.caminho = Path(caminho)
        self.lock = Lock()
        self.wb = self._abrir_ou_criar()

    def _abrir_ou_criar(self):
        if self.caminho.exists():
            workbook = load_workbook(self.caminho)
        else:
            workbook = Workbook()
            workbook.active.title = SHEET_DADOS

        self._garantir_estrutura(workbook)
        self._salvar_workbook(workbook)
        log.info("Planilha pronta: %s", self.caminho)
        return workbook

    def _garantir_estrutura(self, workbook) -> None:
        self._configurar_sheet_dados(workbook)
        self._configurar_sheet_resumo(workbook)

    def _configurar_sheet_dados(self, workbook) -> None:
        if SHEET_DADOS in workbook.sheetnames:
            ws = workbook[SHEET_DADOS]
        else:
            ws = workbook.create_sheet(SHEET_DADOS)

        for indice, coluna in enumerate(COLUNAS, start=1):
            celula = ws.cell(row=1, column=indice, value=coluna["titulo"])
            celula.font = Font(name="Arial", size=11, bold=True, color=COR_CABECALHO_FG)
            celula.fill = PatternFill("solid", start_color=COR_CABECALHO_BG)
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celula.border = borda_fina()
            ws.column_dimensions[get_column_letter(indice)].width = coluna["largura"]

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _configurar_sheet_resumo(self, workbook) -> None:
        if SHEET_RESUMO in workbook.sheetnames:
            ws = workbook[SHEET_RESUMO]
            for intervalo in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(intervalo))
            for linha in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2):
                for celula in linha:
                    celula.value = None
                    celula.fill = PatternFill(fill_type=None)
                    celula.border = Border()
        else:
            ws = workbook.create_sheet(SHEET_RESUMO)

        dados_sheet = f"'{SHEET_DADOS}'"

        ws["A1"] = "Resumo de Chamados"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color=COR_CABECALHO_BG)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:B1")

        metricas = [
            ("Total de Chamados", f"=COUNTA({dados_sheet}!A:A)-1"),
            ("Em Aberto", f'=COUNTIF({dados_sheet}!G:G,"Aberto")'),
            ("Em Andamento", f'=COUNTIF({dados_sheet}!G:G,"Em andamento")'),
            ("Resolvidos", f'=COUNTIF({dados_sheet}!G:G,"Resolvido")'),
            ("Prioridade Alta", f'=COUNTIF({dados_sheet}!F:F,"Alta")'),
            ("Prioridade Média", f'=COUNTIF({dados_sheet}!F:F,"Média")'),
            ("Prioridade Baixa", f'=COUNTIF({dados_sheet}!F:F,"Baixa")'),
            ("Ultima Atualizacao", '=TEXT(NOW(),"DD/MM/YYYY HH:MM")'),
        ]

        for linha, (label, formula) in enumerate(metricas, start=3):
            ws.cell(row=linha, column=1, value=label).font = Font(name="Arial", size=10, bold=True)
            celula = ws.cell(row=linha, column=2, value=formula)
            celula.font = Font(name="Arial", size=10)
            celula.alignment = Alignment(horizontal="center")
            celula.border = borda_fina()

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18

    def _salvar_workbook(self, workbook=None) -> None:
        (workbook or self.wb).save(self.caminho)

    def proximo_sequencial(self) -> int:
        ws = self.wb[SHEET_DADOS]
        if ws.max_row <= 1:
            return 1

        maior = 0
        for linha in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            valor = str(linha[0].value or "")
            try:
                maior = max(maior, int(valor.split("-")[-1]))
            except (ValueError, IndexError):
                continue
        return maior + 1

    def salvar_chamado(self, chamado: dict) -> None:
        with self.lock:
            ws = self.wb[SHEET_DADOS]
            nova_linha = ws.max_row + 1
            usar_fundo_par = nova_linha % 2 == 0
            fundo_par = PatternFill("solid", start_color=COR_LINHA_PAR)

            for indice, coluna in enumerate(COLUNAS, start=1):
                valor = chamado.get(coluna["campo"], "")
                celula = ws.cell(row=nova_linha, column=indice, value=valor)
                celula.font = Font(name="Arial", size=10)
                celula.alignment = Alignment(vertical="center", wrap_text=True)
                celula.border = borda_fina()

                if usar_fundo_par:
                    celula.fill = fundo_par

                if coluna["campo"] == "prioridade":
                    cor = PRIORIDADE_CORES.get(valor)
                    if cor:
                        celula.fill = PatternFill("solid", start_color=cor)
                        celula.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                elif coluna["campo"] == "status":
                    cor = STATUS_CORES.get(valor)
                    if cor:
                        celula.fill = PatternFill("solid", start_color=cor)
                        celula.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

            ws.row_dimensions[nova_linha].height = 35
            ws.auto_filter.ref = ws.dimensions
            self._salvar_workbook()
            log.debug("Chamado salvo na linha %s: %s", nova_linha, chamado.get("id", ""))

    def fechar(self) -> None:
        try:
            self._salvar_workbook()
        except Exception as exc:
            log.warning("Erro ao salvar planilha ao fechar: %s", exc)
